import os
import openpyxl
from openpyxl import load_workbook, Workbook

# Retrieve current working directory (`cwd`)
cwd = os.getcwd()

# Change directory 
os.chdir(cwd)
print(cwd)

wb = Workbook()
wb.save('IDs.xlsx')

wb = load_workbook('IDs.xlsx')
ws = wb.active

filename = input("what is the file name? include .fasta " )
splitter = input("what is the splitter?")



fasta_file = open (filename) 
list_of_ids=[]
for i in fasta_file:
	if ">" in i:
		list_of_ids.append(i)

rownum=1
for i in list_of_ids:
	rownum =rownum + 1
	ID = []	
	ID.append(i.split(splitter))
	for i in ID:
		for col in i:
			columnnum = i.index(col) +1
			ws.cell(row = rownum, column = columnnum).value= col




wb.save('IDs.xlsx')